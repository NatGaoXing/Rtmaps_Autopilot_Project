import rtmaps.types
import numpy as np
import rtmaps.core as rt 
import rtmaps.reading_policy 
from rtmaps.base_component import BaseComponent

from openpyxl import Workbook, load_workbook


class rtmaps_python(BaseComponent):
    def __init__(self):
        BaseComponent.__init__(self)

        self.wb = load_workbook('C:/Users/GAO Xing/Desktop/RtMapsGit/Rtmaps_Autopilot_Project/AMI_Project/data/reference_trajectory_0.xlsx')
        self.ws = self.wb['GPS']
        self.nb_row = self.ws.max_row

    def Dynamic(self):
        self.add_output("reference", rtmaps.types.FLOAT64)

    def Birth(self):
        pass

    def Core(self):
        reference_trajectory = []
        for i in range(1, self.nb_row):
            x = 'A' + str(i)
            y = 'B' + str(i)
            reference_trajectory.append(self.ws[x].value)
            reference_trajectory.append(self.ws[y].value)

        self.outputs["reference"].write(np.array(reference_trajectory))
        # self.outputs["x_ref"].write(reference_trajectory_x)
        # self.outputs["y_ref"].write(reference_trajectory_y)

    def Death(self):
        pass
