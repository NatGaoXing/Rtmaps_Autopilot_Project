import rtmaps.types
import numpy as np
import rtmaps.core as rt 
import rtmaps.reading_policy 
from rtmaps.base_component import BaseComponent

from openpyxl import Workbook, load_workbook
import math


class rtmaps_python(BaseComponent):
    def __init__(self):
        BaseComponent.__init__(self)

        self.wb = load_workbook('C:/Users/GAO Xing/Desktop/RtMapsGit/Rtmaps_Autopilot_Project/AMI_Project/data/record_data.xlsx')
        self.ws = self.wb['GPS']
        nb_row = self.ws.max_row
        self.ws.delete_rows(1, nb_row)

    def Dynamic(self):
        self.add_input("in", rtmaps.types.ANY)
        self.add_output("out", rtmaps.types.AUTO)

    def Birth(self):
        print("Python Birth")

    def Core(self):
        location = self.inputs["in"].ioelt.data

        gps_x = location[0]
        gps_y = location[1]
        gps_point = [gps_x, gps_y]
        last_row_x = 'A' + str(self.ws.max_row)
        last_row_y = 'B' + str(self.ws.max_row)

        if self.ws.max_row < 2:
            self.ws.append(gps_point)
        elif math.sqrt(math.pow((gps_x - self.ws[last_row_x].value), 2) + math.pow((gps_y - self.ws[last_row_y].value), 2)) > 2:
            self.ws.append(gps_point)
        else:
            pass

        self.wb.save('C:/Users/GAO Xing/Desktop/RtMapsGit/Rtmaps_Autopilot_Project/AMI_Project/data/record_data.xlsx')

        self.outputs["out"].write(location)

    def Death(self):
        pass
